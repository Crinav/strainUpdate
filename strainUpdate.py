import re, os, xlrd, requests
from openpyxl import load_workbook
from scrapy.http import HtmlResponse
from lxml import html
from openpyxl.styles import Alignment
from datetime import date 

pathInput = ''
# Recup le path du fichier d'origine 
while True:
    try:
        pathInput = input("Please enter path to your xlsx file (e.g. C:\\path\\to\\my\\file\\name.xlsx)\n").strip()

        if os.path.isfile(pathInput):
            break       
        else:
            print("File doesn't exist on drive")
    except ValueError:
        print("File doesn't exist on drive")
# Le dossier du fichier d'origine
pathBase = os.path.dirname(pathInput)
# le nom du fichier d'origine sans extension
head, tail = os.path.split(pathInput)
fileInput = tail.replace('.xlsx', '')

print("the script could be effective in several minutes ! Be patient !!! ;)")
#création d'un répertoire de travail s'il n'existe pas déjà
if not os.path.exists(pathBase+'/WorkInProgress'):
    os.makedirs(pathBase+'/WorkInProgress')
    # Date du jour
today = date.today()
tempfile =  fileInput+'Updated.'+str(today)+".xlsx"
rowToChange = 0



# Chargement du fichier et de l'onglet pour xlrd
try:
    wb = xlrd.open_workbook(pathInput)
    sheet = wb.sheet_by_index(0)
    wb.release_resources()
    del wb
    wb = load_workbook(pathInput)
except:
    print("Unable to open sheet, please verify first sheet's name is 'strain-1'")
    
# On enregistre une copie du fichier pour travailler dessus
try:
   wb.save(pathBase+"/WorkInProgress/"+tempfile) 
except:
    print(f"Unable to save {tempfile} ")
    
wb.close()

# fonction qui concatène les string d'une list
def concate_list(list):
    result= ''
    for element in list:
        result += str(element)
    return result    


def readExcel(sheet, obj, wb, ws):        
    result = {}
    url = 'http://gcm.wfcc.info/Strain_numberToInfoServlet?strain_number='+obj['ref']
    try:
        mytimeout = (5, 8)
        response = requests.get(url, timeout = mytimeout)
        # Si la requete s'est bien passee
        if response.status_code == 200 :
            response.close()
            # Parsing de la reponse
            root = html.fromstring(response.content)

            # strain = root.xpath('/html/body/div[4]/div[1]/table/tbody/tr/td[contains(., "Strain Number:")]/following-sibling::td/text()')
            # if strain:
            #     result["strain"] = strain[0].strip()

            name =  root.xpath('/html/body/div[4]/div[1]/table/tbody/tr[2]/td[2]/a/strong/i/text()') 
            if name:
                result["name"] = name[0].strip()

            isolated = root.xpath('/html/body/div[4]/div[1]/table/tbody/tr/td[contains(., "Isolated From:")]/following-sibling::td/text()')   
            if isolated:
                result["isolated"] = isolated[0].strip()
                
            # literature = root.xpath('/html/body/div[4]/div[1]/table/tbody/tr/td[contains(., "Literature:")]/following-sibling::td/text()')   
            # if literature:
            #     result["literature"] = literature[0].strip()

            geo = root.xpath('/html/body/div[4]/div[1]/table/tbody/tr/td[contains(., "Geographic Origin:")]/following-sibling::td/text() ')
            if geo:
                result["geo"] = geo[0].strip()

            medium = root.xpath('/html/body/div[4]/div[1]/table/tbody/tr/td[contains(., "Medium Name:")]/following-sibling::td/text()') 
            if medium:
                result["medium"] = medium[0].strip()

            temp = root.xpath('/html/body/div[4]/div[1]/table/tbody/tr/td[contains(., "Optimum Temperature For Growth:")]/following-sibling::td/text()') 
            if temp:
                result["temp"] = temp[0].strip()

            date = root.xpath('/html/body/div[4]/div[1]/table/tbody/tr/td[contains(., "Date of Isolation:")]/following-sibling::td/text()') 
            if date:
                result["date"] = date[0].strip()

            app = root.xpath('/html/body/div[4]/div[1]/table/tbody/tr/td[contains(., "Application:")]/following-sibling::td/text()') 
            if app:
                result["app"] = app[0].strip()

            table = root.xpath('/html/body/div[4]/table[2]/tr/td[@class="ve16t"]/text()')
            if table:
                publi = root.xpath('//*[@id="mainI"]/table[@class="ve16"]/tr/td[@class="ve16t" and contains(.,"Publications")]/../../following-sibling::table[@id="bacteria_static"][1]/tbody/tr/td/table//child::td/a/text()')
                if publi:
                    # Recup des link
                    publiLink = root.xpath('//*[@id="mainI"]/table[@class="ve16"]/tr/td[@class="ve16t" and contains(.,"Publications")]/../../following-sibling::table[@id="bacteria_static"][1]/tbody/tr/td/table//child::td/a/@href')
                    i=0
                    j=1
                    # Nettoyage et ajencement des resultats
                    for val in publiLink:
                        publiLink[i]={"name":publi[j-1]+' '+publi[j].replace('\xa0', ' '),"href":val}
                        i+=1
                        j+=2
                    result["publications"] = publiLink
                
                nodepatent = root.xpath('//*[@id="mainI"]/table[@class="ve16"]/tr/td[@class="ve16t" and contains(.,"Patents")]/../../following-sibling::table[@id="bacteria_static"][1]/tbody/tr/td/table//following-sibling::tr/self::node()')
                patentLink = root.xpath('//*[@id="mainI"]/table[@class="ve16"]/tr/td[@class="ve16t" and contains(.,"Patents")]/../../following-sibling::table[@id="bacteria_static"][1]/tbody/tr/td/table/descendant::td/a/@href')
                j=0
                patentString =[]            
                for node in nodepatent:
                    lst = node.xpath('*//text()')
                    #efface le dernier node
                    if node == nodepatent[len(nodepatent)-1]:
                        del nodepatent[len(nodepatent)-1]
                        break 
                    # Supprime des chiffres superflus
                    if 'PatentNo' in lst:
                        j+=1
                        lst.remove(str(j)) 
                    # Rectifie certaines valeurs manquantes            
                    if len(lst) == 2:
                        patentString.append(concate_list(lst)+'N/A\n')
                    # Enregistrement normal
                    elif len(lst) == 3:
                        patentString.append(concate_list(lst)+'\n')           
                result['patent']= patentString
                result["patentLink"] = patentLink
            if bool(result):
                print(f"Writing result for {obj['ref']}")
                # Ecriture des cells
                writeCell(obj['row'], result, wb, ws)
            else:
                print(f"No result for {obj['ref']}")
        else:
            print("Something went wrong with Global Catalog of Microorganisms...") 
    except:
        print("Something went wrong with internet connection...") 
        
    
                     
            
def writeCell(rowToChange, result, wb, ws):
    # Enregistrement des valeurs dans les cellules
    if rowToChange > 0:
        # Agrandir certaines colonnes
        ws.column_dimensions['BZ'].width = 50
        ws.column_dimensions['CB'].width = 100
        ws.column_dimensions['CA'].width = 100
        if "name" in result:
            cellName = ws.cell(rowToChange,46)
            cellName.value = result["name"]
        if "temp" in result:
            cellTemp = ws.cell(rowToChange,58)
            cellTemp.value = result["temp"]
        if "isolated" in result:
            cellIsolated = ws.cell(rowToChange,37)
            cellIsolated.value = result["isolated"]
        # if "literature" in result:
        #     ws.column_dimensions['CA'].width = 50
        #     cellLiterature = ws.cell(rowToChange,53)
        #     cellLiterature.value = result["literature"]

        # Attention : colonne literature !
        if "publications" in result:
            cellPublications = ws.cell(rowToChange,79)
            cellPublications.alignment = Alignment(wrapText=True)
            dataStr =''

            for val in result["publications"]:
                dataStr += val['name']+'\n'+val['href']+'\n\n'
               
            cellPublications.value = dataStr.strip()
            
        if "date" in result:
            cellDate = ws.cell(rowToChange,36)
            cellDate.value = result["date"]
        if "geo" in result:
            cellGeo = ws.cell(rowToChange,33)
            cellGeo.value = result["geo"]
        if "medium" in result:
            cellMedium = ws.cell(rowToChange,61)
            cellMedium.value = result["medium"]
        if "app" in result:
            cellApp = ws.cell(rowToChange,78)
            cellApp.value = result["app"]
        if "patent" in result:
            cellPatent = ws.cell(rowToChange,80)
            cellPatent.alignment = Alignment(wrapText=True)
            patentStr =''
            i=0
            j=0
            # Ajout des 4 premieresvaleurs + l'url
            for val in result["patent"]:
                if j == 4:
                    patentStr += result["patentLink"][i]+'\n\n'
                    i+=1
                    j = 0                
                else:
                    patentStr += val 
                j+=1
            cellPatent.value = patentStr
    # Sauvegarde du fichier excel      
    wb.save(pathBase+"/WorkInProgress/"+tempfile)


def checkNumbers():
    
    # Chargement de la copie du fichier 
    wb=load_workbook(pathBase+"/WorkInProgress/"+tempfile)
    # déclare l'onglet
    ws=wb["strain-1"]
    # List de toutes les références sur lesquels on va chercher les infos (colonne excel K)
    arrayRef = []
    for cell in ws['K']:
        # On ne prend pas les deux premières row
        if cell.value != None and cell.value != "Equivalent dans une autre collection":
            cell.value = str(cell.value).strip().replace('\n', ',')
            arrayRef.append({"ref" :cell.value.split(',')[0], "row" : cell.row})
        
    
    i=0
    for val in arrayRef:
        i+=1
        print(f"Working on ... {val['ref']} ({i}/{len(arrayRef)} strains)")
        readExcel(sheet, val, wb, ws)
    
checkNumbers()